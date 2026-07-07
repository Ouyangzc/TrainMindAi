"""XLSX parser."""

from openpyxl import load_workbook

from app.models.kb import DocumentPage


def _cell_text(value) -> str:  # noqa: ANN001
    return "" if value is None else str(value)


def _rows_to_markdown(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    width = max(len(row) for row in rows)
    normalized = [row + [""] * (width - len(row)) for row in rows]
    header = normalized[0]
    separator = ["---"] * width
    body_rows = normalized[1:]

    def fmt(row: list[str]) -> str:
        return "| " + " | ".join(row) + " |"

    lines = [fmt(header), fmt(separator)]
    lines.extend(fmt(row) for row in body_rows)
    return "\n".join(lines)


async def parse_xlsx(
    file_path: str, document_id: int, document_version_id: int
) -> list[DocumentPage]:
    """Convert each workbook sheet into a Markdown-table page."""
    wb = load_workbook(file_path, data_only=True, read_only=True)
    pages: list[DocumentPage] = []

    for index, sheet in enumerate(wb.worksheets, start=1):
        rows = [[_cell_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
        text = _rows_to_markdown(rows)
        pages.append(
            DocumentPage(
                document_id=document_id,
                document_version_id=document_version_id,
                page_number=index,
                title=sheet.title,
                text=text,
                metadata_json={"sheet_name": sheet.title},
            )
        )

    return pages
